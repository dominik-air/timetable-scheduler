from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import pandas as pd
import numpy as np
from .data_structures import process_image_manager


def export_matrix_to_excel(matrix: np.ndarray, filename: str = 'ResultSchedule'):
    image_process = process_image_manager.process_image
    writer = pd.ExcelWriter(f'{filename}.xlsx', engine='xlsxwriter')
    start_cols = [0, 15, 30, 45, 60]
    row_indexes = []
    groups = []
    for i in range(5):
        for subgroup in 'AB':
            groups.append(f'grupa {i + 1}{subgroup}')
    for i in range(8, 20):
        for j in range(0, 60, 5):
            if j == 0 or j == 5:
                row_indexes.append(f"{i}:0{j}")
            else:
                row_indexes.append(f"{i}:{j}")
    lectures = {}
    labs = {}
    auditoriums = {}
    for i, day in enumerate(matrix):
        for n, hour in enumerate(day):
            for m, group in enumerate(hour):
                if day[n, m] != 0:
                    course = image_process.courses[day[n, m]]
                    # zapisywanie indeksów macierzy, gdzie zaczynają się zajęcia, ile trwają
                    if 'wyklad' in course.name:
                        length = course.hours_weekly
                        j = round(length * 60 / 5)
                        if course.name not in lectures:
                            lectures[course.name] = (i, n, m, n + j)
                    elif 'laboratoryjne' in course.name:
                        length = course.hours_weekly
                        j = round(length * 60 / 5)
                        if course.name not in labs:
                            labs[course.name] = (i, n, m, n + j)
                    elif 'audytoryjne' in course.name:
                        length = course.hours_weekly
                        j = round(length * 60 / 5)
                        if course.name not in auditoriums:
                            auditoriums[course.name] = (i, n, m, n + j)

        df = pd.DataFrame(day, index=row_indexes, columns=groups)
        # nie wrzucamy zer do excela
        df.replace(0, " ", inplace=True)
        # wszystkie dni tygodnia na jednym arkuszu, dni zaczynaja się od indeksów w liście start_cols
        df.to_excel(writer, sheet_name='Result', startrow=0, startcol=start_cols[i])
    writer.save()

    wb = load_workbook(f'{filename}.xlsx')
    ws1 = wb.active
    # przesuniecie planu 'w dół' żeby zrobić miejsce na nazwy dni tygodnia
    ws1.move_range("A1:BS145", rows=1)
    # scalanie komórek na nazwy dni tygodnia
    ws1.merge_cells("B1:K1")
    ws1.merge_cells("Q1:Z1")
    ws1.merge_cells("AF1:AO1")
    ws1.merge_cells("AU1:BD1")
    ws1.merge_cells("BJ1:BS1")
    currentCell = ws1.cell(row=1, column=2)
    currentCell.value = "Poniedzialek"
    currentCell.alignment = Alignment(horizontal='center')

    currentCell = ws1.cell(row=1, column=17)
    currentCell.value = "Wtorek"
    currentCell.alignment = Alignment(horizontal='center')

    currentCell = ws1.cell(row=1, column=32)
    currentCell.value = "Sroda"
    currentCell.alignment = Alignment(horizontal='center')

    currentCell = ws1.cell(row=1, column=47)
    currentCell.value = "Czwartek"
    currentCell.alignment = Alignment(horizontal='center')

    currentCell = ws1.cell(row=1, column=62)
    currentCell.value = "Piatek"
    currentCell.alignment = Alignment(horizontal='center')
    wb.save(f"{filename}.xlsx")
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )
    #  scalanie komórek, oraz wstawianie odpowiednich nazw, formatowanie
    for wyklad in lectures:
        ws1.merge_cells(start_row=lectures[wyklad][1] + 3, start_column=lectures[wyklad][0] * 15 + 2,
                        end_row=lectures[wyklad][3] + 2, end_column=lectures[wyklad][0] * 15 + 11)
        ws1.cell(row=lectures[wyklad][1] + 3, column=lectures[wyklad][0] * 15 + 2).value = wyklad[:-7]
        ws1.cell(row=lectures[wyklad][1] + 3, column=lectures[wyklad][0] * 15 + 2).alignment = Alignment(
            horizontal='center',
            vertical='center')
        yellow = "00FFFF00"
        ws1.cell(row=lectures[wyklad][1] + 3, column=lectures[wyklad][0] * 15 + 2).fill = PatternFill(
            start_color=yellow, end_color=yellow,
            fill_type="solid")

        for col in range(lectures[wyklad][0] * 15 + 2, lectures[wyklad][0] * 15 + 12):
            ws1.cell(row=lectures[wyklad][1] + 3, column=col).border = thin_border
            ws1.cell(row=lectures[wyklad][3] + 2, column=col).border = thin_border
        for row in range(lectures[wyklad][1] + 3, lectures[wyklad][3] + 3):
            ws1.cell(row=row, column=lectures[wyklad][0] * 15 + 2).border = thin_border
            ws1.cell(row=row, column=lectures[wyklad][0] * 15 + 11).border = thin_border

    for aud in auditoriums:
        start_col = auditoriums[aud][0] * 15 + 2 + auditoriums[aud][2]
        ws1.merge_cells(start_row=auditoriums[aud][1] + 3, start_column=start_col, end_row=auditoriums[aud][3] + 2,
                        end_column=start_col + 1)
        ws1.cell(row=auditoriums[aud][1] + 3, column=start_col).value = aud[:-13]
        ws1.cell(row=auditoriums[aud][1] + 3, column=start_col).alignment = Alignment(horizontal='center',
                                                                                      vertical='center',
                                                                                      text_rotation=180)
        blue = '000000FF'
        ws1.cell(row=auditoriums[aud][1] + 3, column=start_col).fill = PatternFill(
            start_color=blue, end_color=blue,
            fill_type="solid")
        for col in range(start_col, start_col+2):
            ws1.cell(row=auditoriums[aud][1] + 3, column=col).border = thin_border
            ws1.cell(row=auditoriums[aud][3] + 2, column=col).border = thin_border
        for row in range(auditoriums[aud][1] + 3, auditoriums[aud][3] + 3):
            ws1.cell(row=row, column=start_col).border = thin_border
            ws1.cell(row=row, column=start_col+1).border = thin_border

    for lab in labs:
        start_col = labs[lab][0] * 15 + 2 + labs[lab][2]
        ws1.merge_cells(start_row=labs[lab][1] + 3, start_column=start_col, end_row=labs[lab][3] + 2,
                        end_column=start_col)
        cell = ws1.cell(row=labs[lab][1] + 3, column=start_col)
        cell.value = lab[:-16]
        ws1.cell(row=labs[lab][1] + 3, column=start_col).alignment = Alignment(horizontal='center',
                                                                               vertical='center', text_rotation=180)
        green = '0000FF00'
        ws1.cell(row=labs[lab][1] + 3, column=start_col).fill = PatternFill(
            start_color=green, end_color=green,
            fill_type="solid")
        for row in range(labs[lab][1] + 3, labs[lab][3] + 3):
            ws1.cell(row=row, column=start_col).border = thin_border

    # zapis
    wb.save(f"{filename}.xlsx")


def export_availability_to_excel(export_type: str = 'lecturer', id: [int, str] = 0, filename: str = 'lecturer_availability'):
    image_process = process_image_manager.process_image
    writer = pd.ExcelWriter(f'{filename}.xlsx', engine='xlsxwriter')
    row_indexes = []
    for i in range(8, 20):
        for j in range(0, 60, 5):
            if j == 0 or j == 5:
                row_indexes.append(f"{i}:0{j}")
            else:
                row_indexes.append(f"{i}:{j}")
    if 'lecturer' in export_type:
        df = pd.DataFrame(image_process.lecturers[id].availability_matrix, index=row_indexes, columns=["Poniedzialek", "Wtorek", "Sroda", "Czwartek", "Piatek"])
        df.replace(0, " ", inplace=True)
        df.to_excel(writer, sheet_name='Result', startrow=0, startcol=0)
    elif 'room' in export_type:
        df = pd.DataFrame(image_process.rooms[id].availability_matrix, index=row_indexes,
                          columns=["Poniedzialek", "Wtorek", "Sroda", "Czwartek", "Piatek"])
        df.replace(0, " ", inplace=True)
        df.to_excel(writer, sheet_name='Result', startrow=0, startcol=0)
    writer.save()
    wb = load_workbook(f'{filename}.xlsx')
    ws1 = wb.active
    ws1.move_range("A1:F145", rows=1)
    ws1.merge_cells("B1:F1")
    currentCell = ws1.cell(row=1, column=2)
    if export_type == 'lecturer':
        currentCell.value = f'{export_type} {id}'
    else:
        currentCell.value = str(id)
    currentCell.alignment = Alignment(horizontal='center')
    thin_border_1 = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000')
    )

    for row in range(3, len(row_indexes)+3):
        for col in range(2, 7):
            cell_1 = ws1.cell(row=row, column=col)
            if cell_1.value == 1:
                green = "0000FF00"
                cell_1.value = " "
                cell_1.fill = PatternFill(
                    start_color=green, end_color=green,
                    fill_type="solid")
            else:
                red = "00FF0000"
                cell_1.value = " "
                cell_1.fill = PatternFill(
                    start_color=red, end_color=red,
                    fill_type="solid")
            cell_1.border = thin_border_1
    wb.save(f"{filename}.xlsx")
