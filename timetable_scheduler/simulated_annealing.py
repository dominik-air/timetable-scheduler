import random
import math
import cProfile
import numpy as np
from dataclasses import dataclass

from typing import List

from .solution import Solution
from .data_structures import process_image_manager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill
import pandas as pd

# TODO: IMO we can move this class to a separate module or at least its export_matrix_to_excel method
@dataclass
class Results:
    initial_solution_matrix: np.ndarray
    best_solution_matrix: np.ndarray
    initial_cost: float
    best_cost: float
    f_cost_changes: List[float]
    temperature_changes: List[float]

    def export_matrix_to_excel(self, filename: str = 'ResultSchedule'):
        # TODO add borders
        image_process = process_image_manager.process_image
        writer = pd.ExcelWriter(f'{filename}.xlsx', engine='xlsxwriter')
        start_cols = [0, 15, 30, 45, 60]
        row_indexes = []
        groups = []
        for i in range(5):
            for subgroup in 'AB':
                groups.append(f'grupa {i + 1}{subgroup}')
        for i in range(8, 20):
            for j in range(0, 60, 5):
                if j == 0 or j == 5:
                    row_indexes.append(f"{i}:0{j}")
                else:
                    row_indexes.append(f"{i}:{j}")
        lectures = {}
        labs = {}
        auditoriums = {}
        for i, day in enumerate(self.best_solution_matrix):
            for n, hour in enumerate(day):
                for m, group in enumerate(hour):
                    if day[n, m] != 0:
                        course = image_process.courses[day[n, m]]
                        # zapisywanie indeksów macierzy, gdzie zaczynają się zajęcia, ile trwają
                        if 'wyklad' in course.name:
                            length = course.hours_weekly
                            j = round(length * 60 / 5)
                            if course.name not in lectures:
                                lectures[course.name] = (i, n, m, n + j)
                        elif 'laboratoryjne' in course.name:
                            length = course.hours_weekly
                            j = round(length * 60 / 5)
                            if course.name not in labs:
                                labs[course.name] = (i, n, m, n + j, course)
                        elif 'audytoryjne' in course.name:
                            length = course.hours_weekly
                            j = round(length * 60 / 5)
                            if course.name not in auditoriums:
                                auditoriums[course.name] = (i, n, m, n + j)

            df = pd.DataFrame(day, index=row_indexes, columns=groups)
            # nie wrzucamy zer do excela
            df.replace(0, " ", inplace=True)
            # wszystkie dni tygodnia na jednym arkuszu, dni zaczynaja się od indeksów w liście start_cols
            df.to_excel(writer, sheet_name='Result', startrow=0, startcol=start_cols[i])
        writer.save()

        wb = load_workbook(f'{filename}.xlsx')
        ws1 = wb.active
        # przesuniecie planu 'w dół' żeby zrobić miejsce na nazwy dni tygodnia
        ws1.move_range("A1:BS145", rows=1)
        # scalanie komórek na nazwy dni tygodnia
        ws1.merge_cells("B1:K1")
        ws1.merge_cells("Q1:Z1")
        ws1.merge_cells("AF1:AO1")
        ws1.merge_cells("AU1:BD1")
        ws1.merge_cells("BJ1:BS1")
        currentCell = ws1.cell(row=1, column=2)
        currentCell.value = "Poniedzialek"
        currentCell.alignment = Alignment(horizontal='center')

        currentCell = ws1.cell(row=1, column=17)
        currentCell.value = "Wtorek"
        currentCell.alignment = Alignment(horizontal='center')

        currentCell = ws1.cell(row=1, column=32)
        currentCell.value = "Sroda"
        currentCell.alignment = Alignment(horizontal='center')

        currentCell = ws1.cell(row=1, column=47)
        currentCell.value = "Czwartek"
        currentCell.alignment = Alignment(horizontal='center')

        currentCell = ws1.cell(row=1, column=62)
        currentCell.value = "Piatek"
        currentCell.alignment = Alignment(horizontal='center')
        wb.save(f"{filename}.xlsx")

        #  scalanie komórek, oraz wstawianie odpowiednich nazw, formatowanie
        for wyklad in lectures:
            ws1.merge_cells(start_row=lectures[wyklad][1] + 3, start_column=lectures[wyklad][0] * 15 + 2,
                            end_row=lectures[wyklad][3] + 2, end_column=lectures[wyklad][0] * 15 + 11)
            ws1.cell(row=lectures[wyklad][1] + 3, column=lectures[wyklad][0] * 15 + 2).value = wyklad[:-7]
            ws1.cell(row=lectures[wyklad][1] + 3, column=lectures[wyklad][0] * 15 + 2).alignment = Alignment(
                horizontal='center',
                vertical='center')
            yellow = "00FFFF00"
            ws1.cell(row=lectures[wyklad][1] + 3, column=lectures[wyklad][0] * 15 + 2).fill = PatternFill(
                start_color=yellow, end_color=yellow,
                fill_type="solid")

        for aud in auditoriums:
            start_col = auditoriums[aud][0] * 15 + 2 + auditoriums[aud][2]
            ws1.merge_cells(start_row=auditoriums[aud][1] + 3, start_column=start_col, end_row=auditoriums[aud][3] + 2,
                            end_column=start_col + 1)
            ws1.cell(row=auditoriums[aud][1] + 3, column=start_col).value = aud[:-13]
            ws1.cell(row=auditoriums[aud][1] + 3, column=start_col).alignment = Alignment(horizontal='center',
                                                                                          vertical='center',
                                                                                          text_rotation=180)
            blue = '000000FF'
            ws1.cell(row=auditoriums[aud][1] + 3, column=start_col).fill = PatternFill(
                start_color=blue, end_color=blue,
                fill_type="solid")

        for lab in labs:
            start_col = labs[lab][0] * 15 + 2 + labs[lab][2]
            ws1.merge_cells(start_row=labs[lab][1] + 3, start_column=start_col, end_row=labs[lab][3] + 2,
                            end_column=start_col)
            cell = ws1.cell(row=labs[lab][1] + 3, column=start_col)
            cell.value = lab[:-16]
            ws1.cell(row=labs[lab][1] + 3, column=start_col).alignment = Alignment(horizontal='center',
                                                                                   vertical='center', text_rotation=180)
            green = '0000FF00'
            ws1.cell(row=labs[lab][1] + 3, column=start_col).fill = PatternFill(
                start_color=green, end_color=green,
                fill_type="solid")

        # zapis
        wb.save(f"{filename}.xlsx")


def exponential_cooling_schedule(T: int, alpha: float, k: int) -> float:
    return T * (alpha ** k)


def linear_cooling_schedule(T: int, alpha: float, k: int) -> float:
    return T - alpha * k


def logarithmic_cooling_schedule(T: int, alpha: int, k: int) -> float:
    return T / (alpha * math.log(k + 1))


def quadratic_cooling_schedule(T: int, alpha: float, k: int) -> float:
    return T / (1 + alpha * (k ** 2))


def bolzmann_cooling_schedule(T: int, alpha: float, k: int) -> float:
    return T / (1 + math.log(k))


def cauchy_cooling_schedule(T: int, alpha: float, k: int) -> float:
    return T / (1 + k)


def SA(Tmax: int = 20, Tmin: int = 5, kmax: int = 5, alpha: float = 0.99,
       cooling_schedule: callable = exponential_cooling_schedule):
    """Simulated annealing algorithm.

    Args:
        Tmax: initial system temperature.
        Tmin: minimal temperature (lower bound) of the cooling process.
        kmax: number of iterations in a given temperature.
        alpha: cooling process coefficient.
        cooling_schedule: cooling_schedule

    """

    x0 = Solution()
    process_image_copy = process_image_manager.process_image

    x0_cost = x0.cost
    x_best = x0
    f_best = x0_cost

    n_iter = 0
    T = Tmax

    print(f_best)
    f_costs: list[float] = [f_best]
    temperatures: list[float] = [T]

    xc = x0
    while T > Tmin:
        print(T)
        for k in range(kmax):
            xp = xc.from_neighbourhood()
            delta = xp.cost - xc.cost
            if delta <= 0:
                xc = xp
                if xc.cost <= f_best:
                    f_best = xc.cost
                    x_best = xc
                    process_image_copy = process_image_manager.process_image
            else:
                sigma = random.random()
                if sigma < math.exp(-delta / T):
                    xc = xp
            f_costs.append(xp.cost)
        n_iter += 1
        xc = x_best
        process_image_manager.process_image = process_image_copy
        T = cooling_schedule(Tmax, alpha, n_iter)
        temperatures.append(T)

    print(f'Best cost = {f_best}')
    return Results(initial_cost=x0_cost,
                   initial_solution_matrix=x0.matrix,
                   best_cost=f_best,
                   best_solution_matrix=x_best.matrix,
                   f_cost_changes=f_costs,
                   temperature_changes=temperatures)


def test_SA(cooling_schedule=exponential_cooling_schedule):
    SA(cooling_schedule=cooling_schedule)
    process_image_manager.reset_process_image()


if __name__ == '__main__':
    test_SA(cooling_schedule=exponential_cooling_schedule)
